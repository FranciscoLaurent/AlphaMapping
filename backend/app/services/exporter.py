"""
Asset data export service for AlphaMapping.
Supports exporting to CSV, Excel, and JSON formats.
"""
import csv
import json
import io
from typing import List, Any
from datetime import datetime


class AssetExporter:
    """Export assets to various formats."""
    
    # Fields to export
    EXPORT_FIELDS = [
        "id", "ip", "port", "protocol", "domain", "host", 
        "title", "server", "country", "city", "org", "last_seen"
    ]
    
    @staticmethod
    def _asset_to_dict(asset: Any) -> dict:
        """Convert asset object to dictionary."""
        return {
            "id": asset.id,
            "ip": asset.ip,
            "port": asset.port,
            "protocol": asset.protocol or "",
            "domain": asset.domain or "",
            "host": asset.host or "",
            "title": asset.title or "",
            "server": asset.server or "",
            "country": asset.country or "",
            "city": asset.city or "",
            "org": asset.org or "",
            "last_seen": asset.last_seen.isoformat() if asset.last_seen else ""
        }
    
    @classmethod
    def to_csv(cls, assets: List[Any]) -> str:
        """
        Export assets to CSV format.
        
        Args:
            assets: List of Asset objects
            
        Returns:
            CSV string
        """
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=cls.EXPORT_FIELDS)
        writer.writeheader()
        
        for asset in assets:
            writer.writerow(cls._asset_to_dict(asset))
        
        return output.getvalue()
    
    @classmethod
    def to_json(cls, assets: List[Any]) -> str:
        """
        Export assets to JSON format.
        
        Args:
            assets: List of Asset objects
            
        Returns:
            JSON string
        """
        data = [cls._asset_to_dict(asset) for asset in assets]
        return json.dumps(data, ensure_ascii=False, indent=2)
    
    @classmethod
    def to_excel(cls, assets: List[Any]) -> bytes:
        """
        Export assets to Excel format.
        
        Args:
            assets: List of Asset objects
            
        Returns:
            Excel file bytes
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Assets"
        
        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        # Write headers
        for col, field in enumerate(cls.EXPORT_FIELDS, 1):
            cell = ws.cell(row=1, column=col, value=field.upper())
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Write data
        for row, asset in enumerate(assets, 2):
            data = cls._asset_to_dict(asset)
            for col, field in enumerate(cls.EXPORT_FIELDS, 1):
                ws.cell(row=row, column=col, value=data.get(field, ""))
        
        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
    
    @classmethod
    def get_filename(cls, format: str) -> str:
        """Generate export filename with timestamp."""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        extensions = {"csv": "csv", "excel": "xlsx", "json": "json"}
        ext = extensions.get(format, "csv")
        return f"assets_export_{timestamp}.{ext}"
